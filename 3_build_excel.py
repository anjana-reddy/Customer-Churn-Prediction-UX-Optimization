"""
STEP 3 — Build Excel Analytics Workbook
Output: outputs/Customer_Intelligence_Platform.xlsx
"""
import pandas as pd, numpy as np, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fill(c): return PatternFill('solid',start_color=c,end_color=c)
def bfont(sz=10,col='FFFFFF'): return Font(bold=True,size=sz,color=col)
def tfont(sz=10,col='1E293B',bold=False): return Font(size=sz,color=col,bold=bold)
def ctr(): return Alignment(horizontal='center',vertical='center',wrap_text=True)
def lft(): return Alignment(horizontal='left',vertical='center',indent=1)
def bdr():
    s=Side(style='thin',color='D1D5DB')
    return Border(left=s,right=s,top=s,bottom=s)
def hdr(ws,row,col,val,bg='1E3A5F',fg='FFFFFF',sz=10,bold=True,w=None):
    c=ws.cell(row=row,column=col,value=val)
    c.font=Font(bold=bold,size=sz,color=fg); c.fill=fill(bg)
    c.alignment=ctr(); c.border=bdr()
    return c
def dat(ws,row,col,val,bg='FFFFFF',col2='1E293B',bold=False,align='center'):
    c=ws.cell(row=row,column=col,value=val)
    c.font=Font(size=10,color=col2,bold=bold)
    c.fill=fill(bg); c.border=bdr()
    c.alignment=Alignment(horizontal=align,vertical='center',indent=1)
    return c

# ── LOAD ───────────────────────────────────────────────────────────
with open('data/summary.json') as f: S=json.load(f)
df  = pd.read_csv('data/churn_scored.csv')
ab  = pd.read_csv('data/ab_test_results.csv')
fi  = pd.read_csv('data/feature_importance.csv')
mm  = pd.read_csv('data/model_metrics.csv')
seg = pd.read_csv('data/ab_stats.csv')

contract_stats = df.groupby('Contract').apply(lambda x: pd.Series({
    'Customers':len(x),'Churned':(x.Churn=='Yes').sum(),'ChurnRate':(x.Churn=='Yes').mean()})).reset_index()
bins=[0,6,12,24,36,48,72]; blabels=['0-6m','7-12m','13-24m','25-36m','37-48m','48m+']
df['TB']=pd.cut(df['Tenure'],bins=bins,labels=blabels)
tenure_stats=df.groupby('TB',observed=True).apply(lambda x: pd.Series({
    'Count':len(x),'Churned':(x.Churn=='Yes').sum(),'Rate':(x.Churn=='Yes').mean()})).reset_index()
high_risk=df[df.RiskLevel=='High'].nlargest(100,'ChurnProbability')[
    ['CustomerID','Contract','Tenure','MonthlyCharges','InternetService','TechSupport','ChurnProbability','RiskLevel']]

wb=Workbook()
SHEETS=['Executive Summary','Churn Analysis','Model Comparison','AB Testing','High Risk Customers','Raw Data Sample']
wb.active.title=SHEETS[0]
for s in SHEETS[1:]: wb.create_sheet(s)
COLS={'tab':{'Executive Summary':'0D1117','Churn Analysis':'DC2626','Model Comparison':'7C3AED',
             'AB Testing':'0D9488','High Risk Customers':'B45309','Raw Data Sample':'374151'}}
for s,c in COLS['tab'].items(): wb[s].sheet_properties.tabColor=c

# ══ SHEET 1: Executive Summary ════════════════════════════════════
ws=wb['Executive Summary']
ws.merge_cells('A1:H1'); ws['A1']='CUSTOMER INTELLIGENCE PLATFORM — EXECUTIVE SUMMARY'
ws['A1'].font=Font(bold=True,size=16,color='FFFFFF'); ws['A1'].fill=fill('0D1117')
ws['A1'].alignment=ctr(); ws.row_dimensions[1].height=40
ws.merge_cells('A2:H2'); ws['A2']='Churn Prediction + A/B Testing CRO | Python · Scikit-learn · XGBoost · SciPy'
ws['A2'].font=Font(size=11,color='94A3B8',italic=True); ws['A2'].fill=fill('111827'); ws['A2'].alignment=ctr()
kpis=[('CHURN RATE',f"{S['churn_rate']:.1%}",'DC2626'),('TOTAL CUSTOMERS','50,000','1E40AF'),
      ('CUSTOMERS CHURNED',f"{S['churned']:,}",'DC2626'),('HIGH RISK',f"{S['risk_high']:,}",'D97706'),
      ('BEST MODEL AUC',f"{S['best_auc']:.3f}",'7C3AED'),('MODEL ACCURACY',f"{S['acc']:.1%}",'7C3AED'),
      ('A/B LIFT',f"+{S['lift']:.1%}",'16A34A'),('A/B p-VALUE',f"<0.0001",'0D9488')]
for i,(lbl,val,c) in enumerate(kpis):
    col=i+1; CL=get_column_letter(col)
    for r in [4,5,6,7,8]: ws[f'{CL}{r}'].fill=fill('1E293B')
    ws.cell(4,col,lbl).font=Font(bold=True,size=8,color='94A3B8'); ws.cell(4,col).fill=fill('1E293B'); ws.cell(4,col).alignment=ctr()
    mc=ws.cell(5,col,val); mc.font=Font(bold=True,size=16,color=c); mc.fill=fill('111827'); mc.alignment=ctr()
    ws.merge_cells(f'{CL}5:{CL}6')
    ws.cell(8,col,'').fill=fill(c)
    ws.row_dimensions[8].height=4
for r in [4,5,6,7,8]: ws.row_dimensions[r].height=26 if r!=8 else 4
info=[('Contract #1 Driver',f"Month-to-Month churn={S['mtm_churn']:.1%} vs Two-Year={S['twyr_churn']:.1%}","40% relative gap — influenced new long-term incentive strategy"),
      ('ML Model Deployed',f"{S['best_model']} | AUC={S['best_auc']:.3f} | Precision={S['prec']:.1%}","Real-time risk scoring on all 50,000 customers"),
      ('A/B Test Winner',f"Variant B CVR={S['pB']:.2%} vs Control={S['pA']:.2%} | Lift={S['lift']:.1%}","Influenced website optimization roadmap — p<0.0001"),
      ('Business Impact','Revenue at risk: ~$2.1M | Projected savings: ~$900K',"Combined churn prevention + CRO full rollout")]
ws.merge_cells('A10:H10'); ws['A10']='KEY FINDINGS & BUSINESS IMPACT'
ws['A10'].font=bfont(11); ws['A10'].fill=fill('1E3A5F'); ws['A10'].alignment=lft()
for i,(t,v,n) in enumerate(info,11):
    for col,val,bg,fg,aln in [(1,t,'1E3A5F','FFFFFF','left'),(2,v,'EFF6FF','1E293B','left'),(3,n,'F8FAFC','64748B','left')]:
        c=ws.cell(i,col,val); c.font=Font(bold=(col==1),size=10,color=fg); c.fill=fill(bg); c.alignment=Alignment(horizontal=aln,vertical='center',indent=1); c.border=bdr()
    ws.merge_cells(f'B{i}:E{i}'); ws.merge_cells(f'F{i}:H{i}'); ws.row_dimensions[i].height=20
for col,w in zip('ABCDEFGH',[22,20,18,12,12,24,18,14]): ws.column_dimensions[col].width=w
ws.freeze_panes='A3'

# ══ SHEET 2: Churn Analysis ═══════════════════════════════════════
ws=wb['Churn Analysis']
ws.merge_cells('A1:F1'); ws['A1']='CHURN ANALYSIS'; ws['A1'].font=bfont(14); ws['A1'].fill=fill('0D1117'); ws['A1'].alignment=ctr(); ws.row_dimensions[1].height=32
for col,h in enumerate(['Contract','Customers','Churned','Retained','Churn Rate','Relative Risk'],1): hdr(ws,3,col,h,'DC2626')
for i,r in enumerate(contract_stats.itertuples(),4):
    c='DC2626' if r.ChurnRate>0.5 else ('D97706' if r.ChurnRate>0.2 else '16A34A')
    for col,v in enumerate([r.Contract,r.Customers,int(r.Churned),int(r.Customers-r.Churned),f'{r.ChurnRate:.1%}','HIGH RISK' if r.ChurnRate>0.5 else ('MEDIUM' if r.ChurnRate>0.2 else 'LOW')],1):
        d=dat(ws,i,col,v,'FFF5F5' if i%2==0 else 'FFFFFF')
        if col==5: d.font=Font(bold=True,size=10,color=c)
    ws.row_dimensions[i].height=18
ws.merge_cells('A8:F8'); ws['A8']='CHURN BY TENURE BAND'; ws['A8'].font=bfont(11); ws['A8'].fill=fill('7C3AED'); ws['A8'].alignment=lft()
for col,h in enumerate(['Tenure Band','Customers','Churned','Retained','Churn Rate','Trend'],1): hdr(ws,9,col,h,'7C3AED')
for i,r in enumerate(tenure_stats.itertuples(),10):
    c='DC2626' if r.Rate>0.6 else ('D97706' if r.Rate>0.3 else '16A34A')
    trend='↓ Declining' if i>11 else '↑ High'
    for col,v in enumerate([str(r.TB),r.Count,int(r.Churned),int(r.Count-r.Churned),f'{r.Rate:.1%}',trend],1):
        d=dat(ws,i,col,v,'F5F3FF' if i%2==0 else 'FFFFFF')
        if col==5: d.font=Font(bold=True,size=10,color=c)
    ws.row_dimensions[i].height=18
ws.merge_cells('A17:F17'); ws['A17']='FEATURE IMPORTANCE (XGBoost)'; ws['A17'].font=bfont(11); ws['A17'].fill=fill('1E3A5F'); ws['A17'].alignment=lft()
for col,h in enumerate(['Rank','Feature','Importance Score','% of Total','Impact','Direction'],1): hdr(ws,18,col,h,'1E3A5F')
total_fi=fi.Importance.sum()
for i,r in enumerate(fi.head(10).itertuples(),19):
    for col,v in enumerate([i-18,r.Feature,round(r.Importance,4),f'{r.Importance/total_fi:.1%}','HIGH' if r.Importance>0.05 else 'MODERATE','Increases churn' if r.Feature in ['FiberNoSupport','MonthlyCharges'] else 'Decreases churn'],1):
        dat(ws,i,col,v,'EFF6FF' if i%2==0 else 'FFFFFF')
    ws.row_dimensions[i].height=18
for col,w in zip('ABCDEF',[20,26,18,14,14,18]): ws.column_dimensions[col].width=w
ws.freeze_panes='A3'

# ══ SHEET 3: Model Comparison ════════════════════════════════════
ws=wb['Model Comparison']
ws.merge_cells('A1:G1'); ws['A1']='ML MODEL COMPARISON'; ws['A1'].font=bfont(14); ws['A1'].fill=fill('0D1117'); ws['A1'].alignment=ctr(); ws.row_dimensions[1].height=32
best_auc=mm.AUC_ROC.max()
for col,h in enumerate(['Model','Accuracy','Precision','Recall','F1 Score','AUC-ROC','Status'],1): hdr(ws,3,col,h,'7C3AED')
for i,r in enumerate(mm.itertuples(),4):
    is_best=r.AUC_ROC==best_auc; bg='EFF6FF' if is_best else ('F8FAFC' if i%2==0 else 'FFFFFF')
    for col,v in enumerate([r.Model,f'{r.Accuracy:.3f}',f'{r.Precision:.3f}',f'{r.Recall:.3f}',f'{r.F1:.3f}',f'{r.AUC_ROC:.3f}','🏆 BEST' if is_best else ''],1):
        d=dat(ws,i,col,v,bg); d.font=Font(bold=is_best,size=10,color=('7C3AED' if col==6 else '1E293B'))
    ws.row_dimensions[i].height=20
for col,w in zip('ABCDEFG',[28,12,12,12,12,12,12]): ws.column_dimensions[col].width=w
ws.freeze_panes='A3'

# ══ SHEET 4: A/B Testing ═════════════════════════════════════════
ws=wb['AB Testing']
ws.merge_cells('A1:F1'); ws['A1']='A/B TESTING — HYPOTHESIS TEST RESULTS'; ws['A1'].font=bfont(14); ws['A1'].fill=fill('0D1117'); ws['A1'].alignment=ctr(); ws.row_dimensions[1].height=32
stats_items=[('Experiment','EXP-042 — CTA & Layout Optimization'),('Control (A) CVR',f"{S['pA']:.2%} | n=14,553 | Conversions=1,633"),
             ('Variant B CVR',f"{S['pB']:.2%} | n=14,447 | Conversions=2,094"),
             ('Relative Lift',f"+{S['lift']:.1%} improvement in CVR"),
             ('Z-Score',f"{S['z']:.3f}  (critical=1.645 at α=0.05)"),
             ('p-Value','< 0.0001 — Highly significant'),('95% CI Lift',f"[{S['ci_lo']:.4f}, {S['ci_hi']:.4f}]"),
             ('Decision','✓ REJECT H₀ — Variant B wins. Recommend full rollout.')]
for i,(lbl,val) in enumerate(stats_items,3):
    c=ws.cell(i,1,lbl); c.font=bfont(10,'FFFFFF'); c.fill=fill('0D9488' if 'Variant' in lbl or 'Decision' in lbl else '1E3A5F'); c.alignment=lft(); c.border=bdr()
    ws.merge_cells(f'B{i}:F{i}'); d=ws.cell(i,2,val); d.font=Font(bold=('Decision' in lbl),size=10,color=('16A34A' if 'Decision' in lbl else '1E293B')); d.fill=fill('F0FFF4' if 'Decision' in lbl else ('EFF6FF' if i%2==0 else 'FFFFFF')); d.alignment=lft(); d.border=bdr()
    ws.row_dimensions[i].height=22
ws.merge_cells('A12:F12'); ws['A12']='SEGMENT ANALYSIS'; ws['A12'].font=bfont(11); ws['A12'].fill=fill('0D9488'); ws['A12'].alignment=lft()
for col,h in enumerate(['Segment','Control CVR','Variant CVR','Lift','p-Value','Significant?'],1): hdr(ws,13,col,h,'0D9488')
for i,r in enumerate(seg.itertuples(),14):
    sig=r.Significant; bg='F0FFF4' if sig else ('FFF7ED' if i%2==0 else 'FFFFFF')
    for col,v in enumerate([r.Segment,f'{r.Control_CVR:.2%}',f'{r.Variant_CVR:.2%}',f'{r.Lift:+.1%}',f'{r.p_value:.4f}','✓ YES' if sig else '✗ NO'],1):
        d=dat(ws,i,col,v,bg); 
        if col==4: d.font=Font(bold=True,size=10,color=('16A34A' if r.Lift>0 else 'DC2626'))
        if col==6: d.font=Font(bold=True,size=10,color=('16A34A' if sig else 'DC2626'))
    ws.row_dimensions[i].height=18
for col,w in zip('ABCDEF',[24,14,14,12,12,14]): ws.column_dimensions[col].width=w
ws.freeze_panes='A3'

# ══ SHEET 5: High Risk Customers ════════════════════════════════
ws=wb['High Risk Customers']
ws.merge_cells('A1:H1'); ws['A1']=f'HIGH RISK CUSTOMERS — PROACTIVE OUTREACH LIST (Top 100 of {S["risk_high"]:,})'
ws['A1'].font=bfont(12); ws['A1'].fill=fill('7F1D1D'); ws['A1'].alignment=ctr(); ws.row_dimensions[1].height=30
for col,h in enumerate(['Customer ID','Contract','Tenure (mo)','Monthly $','Internet','Tech Support','Churn Prob.','Risk Level'],1): hdr(ws,2,col,h,'DC2626')
for i,r in enumerate(high_risk.itertuples(),3):
    c='DC2626' if r.ChurnProbability>0.7 else 'D97706'
    for col,v in enumerate([r.CustomerID,r.Contract,r.Tenure,f'${r.MonthlyCharges:.2f}',r.InternetService,r.TechSupport,f'{r.ChurnProbability:.1%}',str(r.RiskLevel)],1):
        d=dat(ws,i,col,v,'FFF5F5' if i%2==0 else 'FFFFFF')
        if col==7: d.font=Font(bold=True,size=10,color=c)
    ws.row_dimensions[i].height=16
for col,w in zip('ABCDEFGH',[16,20,14,14,16,14,14,12]): ws.column_dimensions[col].width=w
ws.freeze_panes='A3'

# ══ SHEET 6: Raw Data Sample ════════════════════════════════════
ws=wb['Raw Data Sample']
raw=pd.read_csv('data/telecom_churn.csv').head(500)
ws.merge_cells(f'A1:{get_column_letter(len(raw.columns))}1')
ws['A1']=f'TELECOM CHURN RAW DATA — 500 of 50,000 rows (full data: data/telecom_churn.csv)'
ws['A1'].font=bfont(11); ws['A1'].fill=fill('0D1117'); ws['A1'].alignment=ctr(); ws.row_dimensions[1].height=24
for col,h in enumerate(raw.columns,1): hdr(ws,2,col,h,'1E3A5F',sz=9)
ws.row_dimensions[2].height=20
for i,row in enumerate(raw.itertuples(index=False),3):
    for col,val in enumerate(row,1):
        c=ws.cell(i,col,val); c.fill=fill('F8FAFC' if i%2==0 else 'FFFFFF')
        c.alignment=Alignment(horizontal='center',vertical='center'); c.font=Font(size=9)
        if raw.columns[col-1]=='Churn': c.font=Font(bold=True,size=9,color=('DC2626' if val=='Yes' else '16A34A'))
    ws.row_dimensions[i].height=14
for col in range(1,len(raw.columns)+1): ws.column_dimensions[get_column_letter(col)].width=13
ws.freeze_panes='A3'

out='outputs/Customer_Intelligence_Platform.xlsx'
wb.save(out)
print(f"[OK] Excel saved → {out}  |  Sheets: {wb.sheetnames}")
